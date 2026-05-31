import os
import io
import csv
import json
import re
from typing import List, Dict, Any, Optional
from collections import defaultdict
from statistics import pstdev
from datetime import datetime, timedelta, date

from fastapi import FastAPI, Response, HTTPException, UploadFile, File, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from starlette.status import (
    HTTP_200_OK, 
    HTTP_201_CREATED, 
    HTTP_204_NO_CONTENT, 
    HTTP_409_CONFLICT, 
    HTTP_401_UNAUTHORIZED
)
from jose import jwt, JWTError
from passlib.context import CryptContext
from dotenv import load_dotenv

# Carga de variables de entorno al inicio
load_dotenv()

# ---- Conexiones a la Base de Datos ----
from model.usuario_connection import UsuarioConnection
from model.tipo_usuario_connection import TipoUsuarioConnection
from model.producto_connection import ProductoConnection
from model.marca_connection import MarcaConnection
from model.cliente_connection import ClienteConnection
from model.venta_connection import VentaConnection
from model.forecast_connection import ForecastConnection

# ---- Esquemas de Validación (Pydantic) ----
from schema.usuario_schema import UsuarioSchema
from schema.tipo_usuario_schema import TipoUsuarioSchema
from schema.producto_schema import ProductoSchema
from schema.marca_schema import MarcaSchema
from schema.cliente_schema import ClienteSchema
from schema.venta_schema import VentaSchema
from schema.auth_schema import LoginSchema
from schema.abcxyz_schema import (
    ABCXYZConfigSchema,
    load_config,
    last_12_month_keys,
    abc_label_from_cumshare,
    xyz_label_from_cv,
    month_key_from_date,
    save_config
)
from schema.ml_schema import (
    ForecastRequest,
    ForecastResponseItem,
    ForecastRunInfo,
    ForecastHistoryDetail,
)

# ---- Módulos de Machine Learning ----
from ml.runtime_xgb import build_feature_row, predict_batch, baseline_from_maps
from ml.lag1_postgres import lag1_from_postgres

# --- Inicialización del Estado ---
LAST_ABCXYZ_RESULT: Optional[Dict[str, Any]] = None

# --- Configuración JWT y Seguridad ---
SECRET_KEY = os.getenv("SECRET_KEY", "CAMBIA_ESTA_CLAVE_SUPER_SECRETA")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 120

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def create_access_token(data: dict, expires_delta: timedelta | None = None):
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

# --- AUTORIZACIÓN POR ROLES (ESPAÑOL) ---
ROL_ADMIN = "administrador"
ROL_USER  = "usuario"

def _decode_roles_from_auth(authorization: str | None) -> list[str]:
    """Lee el JWT del header Authorization y devuelve los roles en minúsculas."""
    if not authorization or not authorization.lower().startswith("bearer "):
        return []
    token = authorization.split(" ", 1)[1]
    try:
        claims = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        roles = claims.get("roles") or []
        if isinstance(roles, str):
            roles = [roles]
        return [str(r).lower() for r in roles]
    except JWTError:
        return []

def require_roles(*allowed: str):
    """Dependency para FastAPI. Ej: Depends(require_roles(ROL_ADMIN))."""
    allowed = {a.lower() for a in allowed}
    async def _dep(authorization: str | None = Header(None)):
        roles = set(_decode_roles_from_auth(authorization))
        if not roles & allowed:
            raise HTTPException(status_code=403, detail="No autorizado")
        return True
    return _dep

# --- Configuración FastAPI y CORS ---
app = FastAPI(title="Bee-Organized API Backend")

# Orígenes permitidos extraídos de tu configuración local/producción
origins = [
    "http://localhost:5173",
    os.getenv("FRONTEND_URL", "https://tu-app-en-netlify.netlify.app")
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True if os.getenv("ALLOW_CREDENTIALS") == "true" else False,
)

# Instanciación global de conexiones de infraestructura base
conn = UsuarioConnection()
tipo_conn = TipoUsuarioConnection()
pconn = ProductoConnection()
marca_conn = MarcaConnection()
cliente_conn = ClienteConnection()
venta_conn = VentaConnection()
forecast_conn = ForecastConnection()


# ============================================================================
# --------- ENDPOINTS: USUARIO ---------
# ============================================================================

@app.get("/api/usuario/usuarios", status_code=HTTP_200_OK)
def root():  
    items = []
    for data in conn.read_usuario():
        items.append({
            "id": data[0],
            "usuario": data[1],
            "nombre": data[2],
            "apellido": data[3],
            "correo": data[4],
            "contrasenia": data[5]
        })
    return items

@app.post("/api/usuario/insert", status_code=HTTP_201_CREATED)
def insert_usuario(user_data: UsuarioSchema):
    """
    Crea el usuario, verifica duplicados y le asigna el rol 'usuario' por defecto.
    """
    data = user_data.dict()
    data.pop("id", None)

    # ✅ HU0002-5: validar usuario duplicado
    if conn.get_by_usuario(data["usuario"]):
        raise HTTPException(
            status_code=HTTP_409_CONFLICT,
            detail="El nombre de usuario ya está registrado. Ingrese un usuario diferente."
        )

    # ✅ HU0002-6: validar correo duplicado
    if conn.get_by_correo(data["correo"]):
        raise HTTPException(
            status_code=HTTP_409_CONFLICT,
            detail="El correo electrónico ya está registrado. Ingrese un correo diferente."
        )

    try:
        data["contrasenia"] = pwd_context.hash(data["contrasenia"])
    except Exception:
        pass

    conn.insert_usuario(data)

    created = conn.get_by_usuario(data["usuario"])
    if created:
        nuevo_id = int(created[0])
        try:
            tipo_conn.insert_tipo_usuario({
                "tipo_usuario": ROL_USER,
                "id_usuario":  nuevo_id
            })
        except Exception as e:
            print("WARN: no se pudo asignar rol por defecto:", e)

    return Response(status_code=HTTP_201_CREATED)

@app.get("/api/usuario/{id}", status_code=HTTP_200_OK)
def filtrar_usuario(id: str):
    data = conn.filtrar_usuario(id)
    if data:
        return {
            "id": data[0],
            "usuario": data[1],
            "nombre": data[2],
            "apellido": data[3],
            "correo": data[4],
            "contrasenia": data[5]
        }
    return {"message": "Usuario no encontrado"}

@app.delete("/api/usuario/delete/{id}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def delete_usuario(id: str):
    conn.delete_usuario(id)
    return Response(status_code=HTTP_204_NO_CONTENT)

@app.put("/api/usuario/update/{id}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def update_usuario(user_data: UsuarioSchema, id: str):
    data = user_data.dict()
    data["id"] = id
    conn.update_usuario(data)
    return Response(status_code=HTTP_204_NO_CONTENT)


# ============================================================================
# --------- ENDPOINTS: TIPOS DE USUARIO ---------
# ============================================================================

@app.get("/api/tipo-usuario/listar", status_code=HTTP_200_OK)
def listar_tipo_usuario():
    items = []
    for t in tipo_conn.read_tipo_usuario():
        items.append({
            "id_tipousuario": t[0],
            "tipo_usuario":   t[1],
            "id_usuario":     t[2],
        })
    return items

@app.get("/api/tipo-usuario/{id_tipousuario}", status_code=HTTP_200_OK)
def obtener_tipo_usuario(id_tipousuario: int):
    t = tipo_conn.filtrar_tipo_usuario(id_tipousuario)
    if t:
        return {
            "id_tipousuario": t[0],
            "tipo_usuario":   t[1],
            "id_usuario":     t[2],
        }
    return {"message": "Tipo de usuario no encontrado"}

@app.get("/api/tipo-usuario/usuario/{id_usuario}", status_code=HTTP_200_OK)
def listar_tipo_usuario_por_usuario(id_usuario: int):
    rows = tipo_conn.listar_por_usuario(id_usuario)
    return [
        {"id_tipousuario": r[0], "tipo_usuario": r[1], "id_usuario": r[2]}
        for r in rows
    ]

@app.post("/api/tipo-usuario/insert", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_201_CREATED)
def insertar_tipo_usuario(payload: TipoUsuarioSchema):
    data = payload.dict()
    data.pop("id_tipousuario", None)
    tipo_conn.insert_tipo_usuario(data)
    return Response(status_code=HTTP_201_CREATED)

@app.put("/api/tipo-usuario/update/{id_tipousuario}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def actualizar_tipo_usuario(payload: TipoUsuarioSchema, id_tipousuario: int):
    data = payload.dict()
    data["id_tipousuario"] = id_tipousuario
    tipo_conn.update_tipo_usuario(data)
    return Response(status_code=HTTP_204_NO_CONTENT)

@app.delete("/api/tipo-usuario/delete/{id_tipousuario}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def eliminar_tipo_usuario(id_tipousuario: int):
    tipo_conn.delete_tipo_usuario(id_tipousuario)
    return Response(status_code=HTTP_204_NO_CONTENT)

def get_roles_for_user(user_id: int) -> list[str]:
    rows = tipo_conn.listar_por_usuario(user_id)
    roles = []
    for r in rows:
        rol = (r[1] if isinstance(r, (list, tuple)) else r.get("tipo_usuario")) or ""
        rol = str(rol).strip().lower()
        if rol:
            roles.append(rol)
    return roles or [ROL_USER]


# ============================================================================
# --------- ENDPOINTS: PRODUCTO ---------
# ============================================================================

@app.get("/api/producto/listar", status_code=HTTP_200_OK)
def listar_productos():
    items = []
    for row in pconn.read_producto():
        items.append({
            "id_producto":     row[0],
            "nombre_producto": row[1],
            "id_marca":        row[2],
            "precio_unitario": float(row[3]) if row[3] is not None else None,
            "stock":           row[4],
        })
    return items

@app.get("/api/producto/listar-view", status_code=HTTP_200_OK)
def listar_productos_con_marca():
    items = []
    for row in pconn.read_producto_view():
        items.append({
            "id_producto":     row[0],
            "nombre_producto": row[1],
            "id_marca":        row[2],
            "precio_unitario": float(row[3]) if row[3] is not None else None,
            "stock":           row[4],
            "nombre_marca":    row[5],
        })
    return items

@app.get("/api/producto/{id}", status_code=HTTP_200_OK)
def obtener_producto(id: int):
    row = pconn.filtrar_producto(id)
    if not row:
        return {"message": "Producto no encontrado"}
    return {
        "id_producto":     row[0],
        "nombre_producto": row[1],
        "id_marca":        row[2],
        "precio_unitario": float(row[3]) if row[3] is not None else None,
        "stock":           row[4],
    }

@app.post("/api/producto/insert", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_201_CREATED)
def crear_producto(prod_data: ProductoSchema):
    data = prod_data.dict()
    data.pop("id_producto", None)
    pconn.insert_producto(data)
    return Response(status_code=HTTP_201_CREATED)

@app.put("/api/producto/update/{id}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def actualizar_producto(prod_data: ProductoSchema, id: int):
    data = prod_data.dict()
    data["id_producto"] = id
    pconn.update_producto(data)
    return Response(status_code=HTTP_204_NO_CONTENT)

@app.delete("/api/producto/delete/{id}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def eliminar_producto(id: int):
    pconn.delete_producto(id)
    return Response(status_code=HTTP_204_NO_CONTENT)


# ============================================================================
# --------- ENDPOINTS: MARCA ---------
# ============================================================================

@app.get("/api/marca/listar", status_code=HTTP_200_OK)
def listar_marcas():
    items = []
    for row in marca_conn.read_marca():
        items.append({
            "id_marca": row[0],
            "nombre_marca": row[1],
        })
    return items

@app.get("/api/marca/{id_marca}", status_code=HTTP_200_OK)
def obtener_marca(id_marca: int):
    row = marca_conn.filtrar_marca(id_marca)
    if row:
        return {"id_marca": row[0], "nombre_marca": row[1]}
    return {"message": "Marca no encontrada"}

@app.post("/api/marca/insert", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_201_CREATED)
def insertar_marca(data: MarcaSchema):
    payload = data.dict()
    payload.pop("id_marca", None)
    marca_conn.insert_marca(payload)
    return Response(status_code=HTTP_201_CREATED)

@app.put("/api/marca/update/{id_marca}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def actualizar_marca(data: MarcaSchema, id_marca: int):
    payload = data.dict()
    payload["id_marca"] = id_marca
    marca_conn.update_marca(payload)
    return Response(status_code=HTTP_204_NO_CONTENT)

@app.delete("/api/marca/delete/{id_marca}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def eliminar_marca(id_marca: int):
    try:
        marca_conn.delete_marca(id_marca)
    except Exception as e:
        raise HTTPException(
            status_code=HTTP_409_CONFLICT,
            detail="No se puede eliminar la marca porque está siendo utilizada por productos."
        ) from e
    return Response(status_code=HTTP_204_NO_CONTENT)


# ============================================================================
# --------- ENDPOINTS: CLIENTE ---------
# ============================================================================

@app.get("/api/cliente/listar", status_code=HTTP_200_OK)
def listar_clientes():
    items = []
    for row in cliente_conn.read_cliente():
        items.append({
            "id_cliente": row[0],
            "nombre_empresa": row[1],
            "ruc": row[2],
            "direccion": row[3],
        })
    return items

@app.post("/api/cliente/insert", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_201_CREATED)
def insert_cliente(payload: ClienteSchema):
    data = payload.dict()
    data.pop("id_cliente", None)

    # ✅ HU0020-4: validar RUC único
    if cliente_conn.get_by_ruc(data["ruc"]):
        raise HTTPException(
            status_code=HTTP_409_CONFLICT,
            detail="El RUC ya existe en la aplicación."
        )

    cliente_conn.insert_cliente(data)
    return Response(status_code=HTTP_201_CREATED)

@app.get("/api/cliente/{id_cliente}", status_code=HTTP_200_OK)
def obtener_cliente(id_cliente: int):
    row = cliente_conn.filtrar_cliente(id_cliente)
    if not row:
        return {"message": "Cliente no encontrado"}
    return {
        "id_cliente": row[0],
        "nombre_empresa": row[1],
        "ruc": row[2],
        "direccion": row[3],
    }

@app.delete("/api/cliente/delete/{id_cliente}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def delete_cliente(id_cliente: int):
    cliente_conn.delete_cliente(id_cliente)
    return Response(status_code=HTTP_204_NO_CONTENT)

@app.put("/api/cliente/update/{id_cliente}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def update_cliente(payload: ClienteSchema, id_cliente: int):
    data = payload.dict()
    data["id_cliente"] = id_cliente

    # ✅ HU0022-4: validar RUC único (excluyendo el propio registro)
    existing = cliente_conn.get_by_ruc(data["ruc"])
    if existing and int(existing[0]) != id_cliente:
        raise HTTPException(
            status_code=HTTP_409_CONFLICT,
            detail="El RUC ya existe en la aplicación."
        )

    cliente_conn.update_cliente(data)
    return Response(status_code=HTTP_204_NO_CONTENT)


# ============================================================================
# --------- ENDPOINTS: VENTA ---------
# ============================================================================

@app.get("/api/venta/listar", status_code=HTTP_200_OK)
def listar_ventas():
    items = []
    for r in venta_conn.read_venta():
        items.append({
            "id_venta": r[0],
            "id_producto": r[1],
            "id_cliente": r[2],
            "fecha": r[3].isoformat() if r[3] else None,
            "cantidad": r[4],
            "importe_total": float(r[5]) if r[5] is not None else 0.0,
            "estado": r[6],
        })
    return items

@app.get("/api/venta/{id_venta}", status_code=HTTP_200_OK)
def obtener_venta(id_venta: int):
    r = venta_conn.filtrar_venta(id_venta)
    if r:
        return {
            "id_venta": r[0],
            "id_producto": r[1],
            "id_cliente": r[2],
            "fecha": r[3].isoformat() if r[3] else None,
            "cantidad": r[4],
            "importe_total": float(r[5]) if r[5] is not None else 0.0,
            "estado": r[6],
        }
    return {"message": "Venta no encontrada"}

@app.post("/api/venta/insert", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_201_CREATED)
def insert_venta(v: VentaSchema):
    data = v.model_dump()
    data.pop("id_venta", None)
    venta_conn.insert_venta(data)
    return Response(status_code=HTTP_201_CREATED)

@app.put("/api/venta/update/{id_venta}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def update_venta(v: VentaSchema, id_venta: int):
    data = v.model_dump()
    data["id_venta"] = id_venta
    venta_conn.update_venta(data)
    return Response(status_code=HTTP_204_NO_CONTENT)

@app.delete("/api/venta/delete/{id_venta}", dependencies=[Depends(require_roles(ROL_ADMIN))], status_code=HTTP_204_NO_CONTENT)
def delete_venta(id_venta: int):
    venta_conn.delete_venta(id_venta)
    return Response(status_code=HTTP_204_NO_CONTENT)


# ============================================================================
# --------- ENDPOINTS: AUTH (LOGIN) ---------
# ============================================================================

@app.post("/api/auth/login")
def login(payload: LoginSchema):
    row = conn.get_by_usuario(payload.usuario)
    if not row:
        raise HTTPException(
            status_code=HTTP_401_UNAUTHORIZED,
            detail="El usuario ingresado no existe"
        )

    stored_pwd = row[5]
    provided = payload.contrasenia

    ok = False
    try:
        ok = pwd_context.verify(provided, stored_pwd)
    except Exception:
        ok = False
    if not ok:
        ok = provided == stored_pwd
    if not ok:
        raise HTTPException(
            status_code=HTTP_401_UNAUTHORIZED,
            detail="La contraseña es incorrecta"
        )

    user_id = int(row[0])
    roles = get_roles_for_user(user_id)

    token = create_access_token({"sub": str(user_id), "usuario": row[1], "roles": roles})

    return {
        "token": token, 
        "user": {
            "id":       user_id,
            "usuario":  row[1],
            "nombre":   row[2],
            "apellido": row[3],
            "correo":   row[4],
            "roles":    roles,
        }
    }


# ============================================================================
# --------- ENDPOINTS: DASHBOARD ---------
# ============================================================================

@app.get("/api/venta/listar-view", status_code=HTTP_200_OK)
def listar_ventas_view():
    items = []
    for r in venta_conn.read_venta_view():
        items.append({
            "id_venta": r[0],
            "id_producto": r[1],
            "producto_nombre": r[2],
            "id_cliente": r[3],
            "cliente_nombre": r[4],
            "fecha": r[5].isoformat() if r[5] else None,
            "cantidad": r[6],
            "importe_total": float(r[7]) if r[7] is not None else 0.0,
            "estado": r[8],
        })
    return items


# ============================================================================
# --------- ENDPOINTS: ANALISIS ABC-XYZ ---------
# ============================================================================

@app.get("/api/abcxyz/config", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))], response_model=ABCXYZConfigSchema)
def get_abcxyz_config():
    return load_config()

@app.put("/api/abcxyz/config", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))], response_model=ABCXYZConfigSchema)
def update_abcxyz_config(cfg: ABCXYZConfigSchema):
    if not (0.0 < cfg.a_cut < cfg.b_cut < 1.0):
        raise HTTPException(400, detail="Relación inválida: requiere 0 < A < B < 1.")
    if not (0.0 < cfg.x_cut < cfg.y_cut):
        raise HTTPException(400, detail="Relación inválida: requiere 0 < X < Y.")
    save_config(cfg)
    return cfg

@app.get("/api/abcxyz/precheck")
def abcxyz_precheck():
    cfg = load_config()
    prods = list(ProductoConnection().read_producto())
    ventas = list(VentaConnection().read_venta())

    reasons = []
    if len(prods) == 0:
        reasons.append("No hay productos.")
    if len(ventas) == 0:
        reasons.append("No hay ventas registradas.")
        
    keys = set(last_12_month_keys())
    hay = False
    for v in ventas:
        f = v[3]
        if f and month_key_from_date(f) in keys:
            hay = True
            break
    if not hay:
        reasons.append("No hay ventas en los últimos 12 meses.")

    return {"ok": len(reasons) == 0, "reasons": reasons, "config": cfg.dict()}

@app.post("/api/abcxyz/run", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))])
def abcxyz_run_from_db():
    global LAST_ABCXYZ_RESULT

    cfg = load_config()
    keys = last_12_month_keys()
    kset = set(keys)

    prows = pconn.read_producto()
    prod_map = {r[0]: {"id": r[0], "name": r[1]} for r in prows}
    vrows = vconn.read_venta()

    monthly_qty = defaultdict(lambda: defaultdict(float))
    monthly_amt = defaultdict(lambda: defaultdict(float))

    for r in vrows:
        pid = r[1]
        f = r[3]
        if not f or pid not in prod_map:
            continue
        mk = month_key_from_date(f)
        if mk not in kset:
            continue
        monthly_qty[pid][mk] += float(r[4] or 0.0)
        monthly_amt[pid][mk] += float(r[5] or 0.0)

    rows = []
    total_revenue = 0.0
    for pid, meta in prod_map.items():
        qty_series = [monthly_qty[pid].get(k, 0.0) for k in keys]
        amt_series = [monthly_amt[pid].get(k, 0.0) for k in keys]
        tot_qty = sum(qty_series)
        tot_amt = sum(amt_series)
        total_revenue += tot_amt

        mean = (tot_qty / 12.0)
        sd = pstdev(qty_series) if any(qty_series) else 0.0
        cv = (sd / mean) if mean > 0 else 999.0

        rows.append({
            "id_producto": pid,
            "producto": meta["name"],
            "qty_series": qty_series,
            "amt_series": amt_series,
            "total_qty": tot_qty,
            "total_revenue": tot_amt,
            "cv": cv,
        })

    use_qty_for_abc = (total_revenue == 0.0)
    rows.sort(key=lambda r: r["total_qty"] if use_qty_for_abc else r["total_revenue"], reverse=True)

    cum = 0.0
    base_total = sum(r["total_qty"] for r in rows) if use_qty_for_abc else total_revenue

    res = []
    for r in rows:
        cum += (r["total_qty"] if use_qty_for_abc else r["total_revenue"])
        share = (cum / base_total) if base_total > 0 else 0.0
        abcl = abc_label_from_cumshare(share, cfg.a_cut, cfg.b_cut)
        xyzl = xyz_label_from_cv(r["cv"], cfg.x_cut, cfg.y_cut)
        res.append({
            **r,
            "ABC": abcl,
            "XYZ": xyzl,
            "ABCXYZ": abcl + xyzl
        })

    grid = {a: {x: 0 for x in ["X","Y","Z"]} for a in ["A","B","C"]}
    for r in res:
        grid[r["ABC"]][r["XYZ"]] += 1
    count = len(res)
    perc = {
        a: {x: (grid[a][x] / count * 100 if count else 0) for x in ["X","Y","Z"]}
        for a in ["A","B","C"]
    }

    top = sorted(res, key=lambda r: r["total_qty"], reverse=True)[:3]
    top_series = [{"name": r["producto"], "qty": r["qty_series"]} for r in top]

    payload = {
        "months": keys,
        "rows": res,
        "matrix": {"grid": grid, "percent": perc},
        "totals": {"revenue": total_revenue, "items": count},
        "top_series": top_series,
        "source": "db",
    }

    LAST_ABCXYZ_RESULT = payload
    return payload

@app.get("/api/abcxyz/template")
def abcxyz_template():
    keys = last_12_month_keys()
    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["id_producto", "producto", "marca", *keys])
    w.writerow(["101", "Producto A", "Marca A", *([0] * 12)])
    w.writerow(["102", "Producto B", "Marca B", *([0] * 12)])
    headers = {"Content-Disposition": "attachment; filename=plantilla_abcxyz.csv"}
    return Response(content=output.getvalue(), media_type="text/csv", headers=headers)

@app.post("/api/abcxyz/import", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))])
async def abcxyz_import(file: UploadFile = File(...)):
    global LAST_ABCXYZ_RESULT
    content = await file.read()
    name = (file.filename or "").lower()

    def detect_month_keys_from_header(header_list):
        months = []
        for h in header_list or []:
            if h is None:
                continue
            s = str(h).strip()
            if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", s):
                months.append(s)
        months = sorted(set(months))
        return months[-12:] if len(months) > 12 else months

    series = []
    keys = []

    if name.endswith(".csv"):
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        header = reader.fieldnames or []
        keys = detect_month_keys_from_header(header)
        if not keys:
            keys = last_12_month_keys()
            missing = [k for k in keys if k not in header]
            if missing:
                raise HTTPException(400, detail=f"No se detectaron columnas YYYY-MM. Faltan: {', '.join(missing)}")

        if len(keys) != 12:
            raise HTTPException(400, detail=f"Se requieren 12 meses. Detectados: {len(keys)}")

        for row in reader:
            series.append({
                "id_producto": int(row.get("id_producto", 0)) if row.get("id_producto") else None,
                "producto": (row.get("producto") or "").strip(),
                "marca": (row.get("marca") or "").strip(),
                "qty_series": [float((row.get(k) or 0)) for k in keys],
            })
    else:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            raise HTTPException(400, detail="Para XLSX instala openpyxl o sube CSV.")

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        header = [str(c).strip() if c is not None else "" for c in header_row]

        keys = detect_month_keys_from_header(header)
        if not keys:
            keys = last_12_month_keys()
            missing = [k for k in keys if k not in header]
            if missing:
                raise HTTPException(400, detail=f"No se detectaron columnas YYYY-MM. Faltan: {', '.join(missing)}")

        if len(keys) != 12:
            raise HTTPException(400, detail=f"Se requieren 12 meses. Detectados: {len(keys)}")

        try:
            idx_pid = header.index("id_producto")
            idx_prod = header.index("producto")
        except ValueError:
            raise HTTPException(400, detail="Encabezados requeridos: id_producto, producto.")

        idx_marca = header.index("marca") if "marca" in header else None
        idx_cols = [header.index(k) for k in keys]

        for r in ws.iter_rows(min_row=2, values_only=True):
            series.append({
                "id_producto": int(r[idx_pid]) if r[idx_pid] not in (None, "") else None,
                "producto": str(r[idx_prod]) if r[idx_prod] is not None else "",
                "marca": str(r[idx_marca]).strip() if idx_marca is not None and r[idx_marca] is not None else "",
                "qty_series": [float(r[i] or 0) for i in idx_cols],
            })

    cfg = load_config()
    res = []
    for s in series:
        tot_qty = float(sum(s["qty_series"]))
        mean = (tot_qty / 12.0)
        sd = pstdev(s["qty_series"]) if any(s["qty_series"]) else 0.0
        cv = (sd / mean) if mean > 0 else 999.0
        res.append({
            **s,
            "amt_series": s["qty_series"],
            "total_qty": tot_qty,
            "total_revenue": tot_qty,
            "cv": cv
        })

    total_base = sum(r["total_revenue"] for r in res)
    use_qty_for_abc = (total_base == 0.0)
    res.sort(key=lambda r: r["total_qty"] if use_qty_for_abc else r["total_revenue"], reverse=True)

    cum = 0.0
    base_total = sum(r["total_qty"] for r in res) if use_qty_for_abc else total_base

    out = []
    for r in res:
        cum += (r["total_qty"] if use_qty_for_abc else r["total_revenue"])
        share = (cum / base_total) if base_total > 0 else 0.0
        abcl = abc_label_from_cumshare(share, cfg.a_cut, cfg.b_cut)
        xyzl = xyz_label_from_cv(r["cv"], cfg.x_cut, cfg.y_cut)
        out.append({**r, "ABC": abcl, "XYZ": xyzl, "ABCXYZ": abcl + xyzl})

    grid = {a: {x: 0 for x in ["X", "Y", "Z"]} for a in ["A", "B", "C"]}
    for r in out:
        grid[r["ABC"]][r["XYZ"]] += 1
    count = len(out)
    perc = {a: {x: (grid[a][x] / count * 100 if count else 0) for x in ["X", "Y", "Z"]} for a in ["A", "B", "C"]}

    top_series = [{"name": r["producto"], "qty": r["qty_series"]} for r in sorted(out, key=lambda r: r["total_qty"], reverse=True)[:3]]

    payload = {
        "months": keys,
        "rows": out,
        "matrix": {"grid": grid, "percent": perc},
        "totals": {"revenue": total_base, "items": count},
        "top_series": top_series,
        "source": "excel",
    }

    LAST_ABCXYZ_RESULT = payload
    return payload

@app.get("/api/abcxyz/last", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))])
def abcxyz_last(source: Optional[str] = None):
    if LAST_ABCXYZ_RESULT is None:
        raise HTTPException(404, detail="Aún no se ha ejecutado ningún análisis ABC-XYZ.")

    if source is not None and LAST_ABCXYZ_RESULT.get("source") != source:
        raise HTTPException(
            404,
            detail=f"No hay un análisis ABC-XYZ reciente con fuente '{source}'. Última fuente: '{LAST_ABCXYZ_RESULT.get('source')}'."
        )
    return LAST_ABCXYZ_RESULT


# ============================================================================
# --------- ENDPOINTS: FORECAST (XGBoost / Baseline) ---------
# ============================================================================

def _last_nonzero(values):
    if not isinstance(values, (list, tuple)):
        return 0.0
    for v in reversed(values):
        try:
            x = float(v or 0)
        except Exception:
            x = 0.0
        if x > 0:
            return x
    return 0.0

def _moving_avg_last_k_nonzero(values, k=3):
    nz = []
    if isinstance(values, (list, tuple)):
        for v in values:
            try:
                x = float(v or 0)
            except Exception:
                x = 0.0
            if x > 0:
                nz.append(x)
    if not nz:
        return 0.0
    k = min(k, len(nz))
    return sum(nz[-k:]) / k

def _to_ym(s) -> str:
    if s is None:
        return ""
    if isinstance(s, (datetime, )):
        return f"{s.year:04d}-{s.month:02d}"
    try:
        txt = str(s)
        if len(txt) >= 7 and txt[4] == '-' and txt[6].isdigit():
            return txt[:7]
    except Exception:
        pass
    return ""

def _ym_to_int(ym: str) -> int:
    return int(ym[:4]) * 12 + int(ym[5:7])

def _avg_last_k_deltas(values, k=3):
    if not isinstance(values, (list, tuple)) or len(values) < 2:
        return 0.0
    deltas = []
    prev = None
    for v in values:
        try:
            x = float(v or 0)
        except Exception:
            x = 0.0
        if prev is not None and x > 0 and prev > 0:
            deltas.append(x - prev)
        prev = x
    if not deltas:
        return 0.0
    k = min(k, len(deltas))
    return sum(deltas[-k:]) / k

def _value_for_month_from_series(months: list[str], serie: list[float], ym: str) -> float | None:
    if not months or len(months) != len(serie):
        return None
    try:
        i = months.index(ym)
        return float(serie[i] or 0.0)
    except (ValueError, Exception):
        return None

def _extrapolate_with_trend(months: list[str], serie: list[float], ym: str) -> float:
    if not months or len(months) != len(serie):
        b = _last_nonzero(serie)
        if b == 0.0:
            b = _moving_avg_last_k_nonzero(serie, k=3)
        if b == 0.0:
            b = sum([float(v or 0) for v in serie]) / 12.0 if serie else 0.0
        return max(0.0, float(b))

    ym_min, ym_max = months[0], months[-1]
    t_min, t_max, t = _ym_to_int(ym_min), _ym_to_int(ym_max), _ym_to_int(ym)

    if t > t_max:
        anchor = _last_nonzero(serie) or float(serie[-1] or 0.0)
        slope = _avg_last_k_deltas(serie, k=3)
        return max(0.0, float(anchor + (t - t_max) * slope))

    if t < t_min:
        deltas = []
        prev = None
        for v in serie:
            x = float(v or 0.0)
            if prev is not None and x > 0 and prev > 0:
                deltas.append(x - prev)
            prev = x
            if len(deltas) >= 3:
                break
        slope = sum(deltas) / len(deltas) if deltas else 0.0
        anchor = float(serie[0] or 0.0)
        return max(0.0, float(anchor - (t_min - t) * slope))

    return max(0.0, float(_last_nonzero(serie)))

def _csv_baseline_for_item(it, target_month: str) -> float:
    if LAST_ABCXYZ_RESULT is None:
        return 0.0

    rows = LAST_ABCXYZ_RESULT.get("rows") or []
    months = LAST_ABCXYZ_RESULT.get("months") or []
    ym = _to_ym(target_month)

    pid = getattr(it, "id_producto", None)
    nombre = (getattr(it, "producto", "") or "").strip().lower()
    cand = None
    for r in rows:
        if pid not in (None, 0) and r.get("id_producto") == pid:
            cand = r
            break
        if cand is None and nombre and (r.get("producto") or "").strip().lower() == nombre:
            cand = r
    if not cand:
        return 0.0

    serie = cand.get("amt_series") or cand.get("qty_series") or []

    val_directo = _value_for_month_from_series(months, serie, ym)
    if val_directo is not None and val_directo > 0:
        return float(val_directo)

    val_trend = _extrapolate_with_trend(months, serie, ym)
    if val_trend > 0:
        return float(val_trend)

    b = _last_nonzero(serie) or _moving_avg_last_k_nonzero(serie, k=3)
    if b == 0.0:
        total = float(cand.get("total_revenue") or cand.get("total_qty") or 0.0)
        b = (total / 12.0) if total > 0 else 0.0
    return float(b)

@app.post("/api/forecast/xgb", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))], response_model=List[ForecastResponseItem])
def forecast_xgb(payload: ForecastRequest):
    rows_features = []
    baselines: List[float] = []
    meses = set()

    for it in payload.items:
        meses.add(it.fecha_mes)

        if payload.origen == "abcxyz_db":
            rows_features.append(
                build_feature_row(
                    producto=it.producto,
                    marca=it.marca,
                    fecha_mes=it.fecha_mes,
                    pct_chg_1=it.pct_chg_1 or 0.0,
                )
            )
            b = lag1_from_postgres(it.id_producto, it.fecha_mes) if it.id_producto else 0.0
            baselines.append(float(b))
        else:
            baselines.append(float(_csv_baseline_for_item(it, it.fecha_mes)))

    if payload.origen == "abcxyz_db":
        y_pred = predict_batch(rows_features, baselines)
    else:
        y_pred = baselines

    periodo_inicio = min(meses) if meses else date.today()
    periodo_fin = max(meses) if meses else date.today()
    horizonte_meses = len(meses)

    run_data = {
        "id_usuario": None,
        "origen": payload.origen or "desconocido",
        "modelo": "xgboost" if payload.origen == "abcxyz_db" else "baseline_csv",
        "modelo_version": "XGB_2025-11-04_v1",
        "periodo_inicio": periodo_inicio,
        "periodo_fin": periodo_fin,
        "horizonte_meses": horizonte_meses,
    }
    id_run = forecast_conn.insert_run(run_data)

    detalles = []
    response_items: List[ForecastResponseItem] = []

    for it, y, b in zip(payload.items, y_pred, baselines):
        cat_abc = it.categoria_abc
        cat_xyz = it.categoria_xyz
        cat_abcxyz = f"{cat_abc}{cat_xyz}" if cat_abc and cat_xyz else None
        id_prod_det = (it.id_producto if it.id_producto not in (0, None) else None) if payload.origen == "abcxyz_db" else None

        detalles.append({
            "id_run": id_run,
            "id_producto": id_prod_det,
            "fecha_mes": it.fecha_mes,
            "venta_predicha": float(y),
            "baseline": float(b),
            "categoria_abc": cat_abc,
            "categoria_xyz": cat_xyz,
            "categoria_abcxyz": cat_abcxyz,
        })

        response_items.append(
            ForecastResponseItem(
                id_producto=it.id_producto or 0,
                producto=it.producto,
                fecha_mes=it.fecha_mes,
                prediccion=float(y),
            )
        )

    forecast_conn.insert_detalle_many(detalles)
    return response_items

@app.get("/api/forecast/history", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))], response_model=List[ForecastRunInfo])
def forecast_history():
    rows = forecast_conn.read_runs()
    out: List[ForecastRunInfo] = []
    for r in rows:
        out.append(
            ForecastRunInfo(
                id_run=r[0], creado_en=r[1], origen=r[2], modelo=r[3],
                modelo_version=r[4], periodo_inicio=r[5], periodo_fin=r[6], horizonte_meses=r[7],
            )
        )
    return out

@app.get("/api/forecast/history/{id_run}", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))], response_model=List[ForecastHistoryDetail])
def forecast_history_detail(id_run: int):
    rows = forecast_conn.read_detalle_by_run(id_run)
    out: List[ForecastHistoryDetail] = []
    for r in rows:
        out.append(
            ForecastHistoryDetail(
                id_detalle=r[0], id_run=r[1], id_producto=r[2], producto=r[3],
                fecha_mes=r[4], venta_predicha=float(r[5]), baseline=float(r[6]),
                categoria_abc=r[7], categoria_xyz=r[8], categoria_abcxyz=r[9],
            )
        )
    return out

@app.delete("/api/forecast/history/{id_run}", dependencies=[Depends(require_roles(ROL_ADMIN, ROL_USER))])
def delete_forecast_run(id_run: int):
    forecast_conn.delete_run(id_run)
    return JSONResponse(content={"ok": True, "id_run": id_run}, status_code=200)